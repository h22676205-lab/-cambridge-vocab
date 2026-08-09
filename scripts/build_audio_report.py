"""
重新產生「單字錄音對照表.xlsx」。

這個腳本直接從 index.html 讀取 CATEGORIES 和 AUDIO_MAP（不是自己維護一份複本），
並比對 audio/ 資料夾裡實際存在的檔案，所以只要 index.html 或 audio/ 資料夾有更新，
重新執行這個腳本就會自動反映最新狀態，不會跟 App 內容漂移不同步。

用法（在 英語單字學習 資料夾內執行）：
    python scripts/build_audio_report.py
"""
import json
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HTML_PATH = os.path.join(BASE_DIR, "index.html")
AUDIO_DIR = os.path.join(BASE_DIR, "audio")
OUT_PATH = os.path.join(BASE_DIR, "單字錄音對照表.xlsx")

FONT = "Arial"


def extract_json_array(html, var_name):
    marker = f"const {var_name} = "
    start = html.index(marker) + len(marker)
    depth = 0
    for i in range(start, len(html)):
        ch = html[i]
        if ch == "[" or ch == "{":
            depth += 1
        elif ch == "]" or ch == "}":
            depth -= 1
            if depth == 0:
                raw = html[start:i + 1]
                # JS object/array literals allow a trailing comma before a closing
                # bracket; plain JSON does not, so strip it before parsing.
                raw = re.sub(r",\s*([}\]])", r"\1", raw)
                return json.loads(raw)
    raise ValueError(f"could not find end of {var_name}")


def main():
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    categories = extract_json_array(html, "CATEGORIES")
    audio_map = extract_json_array(html, "AUDIO_MAP")

    have_files = set(f[:-4] for f in os.listdir(AUDIO_DIR) if f.endswith(".mp3"))

    rows = []
    for cat, words in categories:
        for en, zh in words:
            spec = audio_map.get(en)
            stems = spec if isinstance(spec, list) else ([spec] if spec else [])
            present = [s[:-4] for s in stems if s[:-4] in have_files]
            has_recording = len(stems) > 0 and len(present) == len(stems)
            filenames = ", ".join(stems) if has_recording else ""
            rows.append({
                "cat": cat, "en": en, "zh": zh,
                "has": "有老師錄音" if has_recording else "沒有錄音（用機器語音）",
                "file": filenames,
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "單字錄音對照表"

    headers = ["分類", "英文單字", "中文", "是否有老師錄音", "音檔檔名"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(name=FONT, color="006100")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(name=FONT, color="9C0006")

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["cat"]).font = Font(name=FONT)
        ws.cell(row=r, column=2, value=row["en"]).font = Font(name=FONT)
        ws.cell(row=r, column=3, value=row["zh"]).font = Font(name=FONT)
        status_cell = ws.cell(row=r, column=4, value=row["has"])
        if row["has"] == "有老師錄音":
            status_cell.fill = green_fill
            status_cell.font = green_font
        else:
            status_cell.fill = red_fill
            status_cell.font = red_font
        ws.cell(row=r, column=5, value=row["file"]).font = Font(name=FONT)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
        r += 1

    # Merge category column visually per contiguous block
    run_start_idx = 0
    for i in range(1, len(rows) + 1):
        at_end = i == len(rows)
        if at_end or rows[i]["cat"] != rows[run_start_idx]["cat"]:
            sheet_start = run_start_idx + 2
            sheet_end = i - 1 + 2
            if sheet_end > sheet_start:
                ws.merge_cells(start_row=sheet_start, start_column=1, end_row=sheet_end, end_column=1)
                ws.cell(row=sheet_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
            run_start_idx = i

    # Summary sheet
    ws2 = wb.create_sheet("統計摘要")
    for c, h in enumerate(["分類", "總單字數", "有錄音", "沒有錄音"], start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cats_in_order = [c for c, _ in categories]
    r2 = 2
    for cat in cats_in_order:
        total_n = sum(1 for row in rows if row["cat"] == cat)
        has_n = sum(1 for row in rows if row["cat"] == cat and row["has"] == "有老師錄音")
        ws2.cell(row=r2, column=1, value=cat).font = Font(name=FONT)
        ws2.cell(row=r2, column=2, value=total_n).font = Font(name=FONT)
        ws2.cell(row=r2, column=3, value=has_n).font = Font(name=FONT)
        ws2.cell(row=r2, column=4, value=total_n - has_n).font = Font(name=FONT)
        r2 += 1

    total_row = r2
    total_all = len(rows)
    has_all = sum(1 for row in rows if row["has"] == "有老師錄音")
    ws2.cell(row=total_row, column=1, value="總計").font = Font(name=FONT, bold=True)
    ws2.cell(row=total_row, column=2, value=total_all).font = Font(name=FONT, bold=True)
    ws2.cell(row=total_row, column=3, value=has_all).font = Font(name=FONT, bold=True)
    ws2.cell(row=total_row, column=4, value=total_all - has_all).font = Font(name=FONT, bold=True)
    ws2.cell(row=total_row + 2, column=1,
             value="備註：同一個英文單字在不同分類重複出現時（例如 fish、orange），共用同一份錄音會一併計入「有錄音」。").font = \
        Font(name=FONT, italic=True, size=9, color="808080")

    for col, width in zip("ABCDE", [22, 16, 10, 22, 20]):
        ws.column_dimensions[col].width = width
    for col, width in zip("ABCD", [24, 12, 10, 10]):
        ws2.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"saved {OUT_PATH}")
    print(f"total rows: {total_all}, has recording: {has_all}, no recording: {total_all - has_all}")


if __name__ == "__main__":
    main()
